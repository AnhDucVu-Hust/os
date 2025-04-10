import sys
import os
import calendar
import numpy
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QLabel, 
                            QPushButton, QFileDialog, QMessageBox, QHBoxLayout, QDateEdit)
from PyQt5.QtCore import Qt, QDate
import pandas as pd
import numpy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from styleframe import StyleFrame, Styler, utils
import roman
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings("ignore")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("OS Document Processor")
        self.setMinimumSize(400, 300)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Date selection section
        date_widget = QWidget()
        date_layout = QVBoxLayout(date_widget)
        
        # From date
        from_date_layout = QHBoxLayout()
        from_date_label = QLabel("Từ ngày:")
        self.from_date = QDateEdit()
        self.from_date.setDate(QDate.currentDate())
        self.from_date.setCalendarPopup(True)
        from_date_layout.addWidget(from_date_label)
        from_date_layout.addWidget(self.from_date)
        date_layout.addLayout(from_date_layout)

        # To date
        to_date_layout = QHBoxLayout()
        to_date_label = QLabel("Đến ngày:")
        self.to_date = QDateEdit()
        self.to_date.setDate(QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        to_date_layout.addWidget(to_date_label)
        to_date_layout.addWidget(self.to_date)
        date_layout.addLayout(to_date_layout)

        layout.addWidget(date_widget)

        # File selection button
        self.select_button = QPushButton("Chọn file Excel")
        self.select_button.clicked.connect(self.select_file)
        layout.addWidget(self.select_button)

        # Status label
        self.status_label = QLabel("Hãy chọn file Excel để xử lý")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)

    def append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None, startcol=0,
                       truncate_sheet=False, **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
        """
        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = StyleFrame.ExcelWriter(filename, mode='a', if_sheet_exists='overlay')

        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)

        # save the workbook
        writer.save()
    
    def find_first_monday(self, year, month):
        d = datetime(year, int(month), 7)
        offset = -d.weekday()  # weekday = 0 means monday
        time = str(d + timedelta(offset))
        y, m, d = str(time.split()[0]).split('-')
        return d+"/"+m+"/"+y
    
    def last_business_day_in_month(self, year, month):
        return str(max(calendar.monthcalendar(year, month)[-1][:5]))+"/"+str(month)+"/"+str(year)

    def get_time(self):
        from_date = self.from_date.date().toString("dd/MM/yyyy")
        to_date = self.to_date.date().toString("dd/MM/yyyy")
        return f"{from_date} đến {to_date}"

    def create_output_folder(self, input_file_path):
        # Lấy thư mục của file input
        input_dir = os.path.dirname(input_file_path)
        if not input_dir:  # Nếu file ở thư mục hiện tại
            input_dir = os.getcwd()
            
        # Tạo tên folder dựa trên ngày
        folder_name = f"OS_Document_{self.from_date.date().toString('dd_MM_yyyy')}_to_{self.to_date.date().toString('dd_MM_yyyy')}"
        output_dir = os.path.join(input_dir, folder_name)
        
        # Tạo folder nếu chưa tồn tại
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        return output_dir

    def select_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Chọn file Excel", "", "Excel Files (*.xlsx)"
        )
        if file_name:
            self.process_file(file_name)

    def process_file(self, file_path):
        try:
            # Kiểm tra ngày
            if self.from_date.date() > self.to_date.date():
                QMessageBox.warning(self, "Cảnh báo", "Ngày bắt đầu không thể sau ngày kết thúc")
                return

            self.status_label.setText("Đang xử lý...")
            QApplication.processEvents()

            # Tạo thư mục output
            output_dir = self.create_output_folder(file_path)

            # Đọc file Excel - dùng cách đọc giống với os_check.py
            df = StyleFrame.read_excel(file_path, sheet_name="Sheet0")
            hop_dong = list(set(df["Hợp đồng"]))
            df_hd = {}
            df_chuan = {}
            
            for hd in hop_dong:
                self.status_label.setText(f"Đang xử lý hợp đồng: {hd}")
                QApplication.processEvents()
                
                df_chuan[hd] = pd.DataFrame(columns=['Id','Row label','Bảo trì','Nâng cấp','Total'])
                df_hd[hd] = df.loc[df["Hợp đồng"]==hd]
                he_thongs = list(set(df_hd[hd]["Hệ thống&CTKT"]))
                id_he_thong = 1
                tong_bao_tri = 0
                tong_nang_cap = 0
                
                for he_thong in he_thongs:
                    df_he_thong = df_hd[hd].loc[df_hd[hd]["Hệ thống&CTKT"]==he_thong]
                    stories = list(set(df_he_thong["Tên story"]))
                    id_story = 1
                    
                    bao_tri = df_he_thong.loc[df_he_thong["Phân loại"]=='Bảo trì']["ULNL task"].sum()
                    nang_cap = df_he_thong.loc[df_he_thong["Phân loại"]=='Nâng cấp']["ULNL task"].sum()
                    total = bao_tri + nang_cap
                    tong_nang_cap += nang_cap
                    tong_bao_tri += bao_tri
                    
                    # Sử dụng pd.concat thay cho append
                    new_row = pd.DataFrame([{
                        'Id': roman.toRoman(id_he_thong),
                        'Row label': he_thong,
                        'Bảo trì': bao_tri,
                        "Nâng cấp": nang_cap,
                        'Total': total
                    }])
                    df_chuan[hd] = pd.concat([df_chuan[hd], new_row], ignore_index=True)
                    
                    for story in stories:
                        df_xet = df_he_thong.loc[df_he_thong["Tên story"]==story]
                        bao_tri = df_xet.loc[df_xet["Phân loại"] == 'Bảo trì']["ULNL task"].sum()
                        nang_cap = df_xet.loc[df_xet["Phân loại"] == 'Nâng cấp']["ULNL task"].sum()
                        df_xet = df_xet[["Hệ thống&CTKT","Tên story","Summary","Phân loại","ULNL task"]]
                        total = bao_tri + nang_cap
                        
                        # Sử dụng pd.concat thay cho append
                        new_row = pd.DataFrame([{
                            'Id': str(roman.toRoman(id_he_thong))+"."+str(id_story),
                            "Row label": story,
                            'Bảo trì': bao_tri,
                            "Nâng cấp": nang_cap,
                            'Total': total
                        }])
                        df_chuan[hd] = pd.concat([df_chuan[hd], new_row], ignore_index=True)
                        
                        tasks = list(set(df_xet["Summary"]))
                        id_task = 1
                        for task in tasks:
                            df_xet2 = df_xet.loc[df_xet["Summary"]==task]
                            bao_tri = df_xet2.loc[df_xet2["Phân loại"] == 'Bảo trì']["ULNL task"].sum()
                            nang_cap = df_xet2.loc[df_xet2["Phân loại"] == 'Nâng cấp']["ULNL task"].sum()
                            total = bao_tri + nang_cap
                            
                            # Sử dụng pd.concat thay cho append
                            new_row = pd.DataFrame([{
                                'Id': id_task,
                                "Row label": task,
                                'Bảo trì': bao_tri,
                                "Nâng cấp": nang_cap,
                                'Total': total
                            }])
                            df_chuan[hd] = pd.concat([df_chuan[hd], new_row], ignore_index=True)
                            id_task += 1
                        id_story += 1
                    id_he_thong += 1
                
                # Add total row with pd.concat
                new_row = pd.DataFrame([{
                    'Id': '',
                    "Row label": 'Tổng',
                    'Bảo trì': tong_bao_tri,
                    "Nâng cấp": tong_nang_cap,
                    'Total': tong_bao_tri + tong_nang_cap
                }])
                df_chuan[hd] = pd.concat([df_chuan[hd], new_row], ignore_index=True)

                # Format output
                df_bbnv = df_chuan[hd].rename(columns={
                    'Id': 'TT',
                    'Row label': 'Nội dung công việc chi tiết',
                    'Nâng cấp': 'Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)',
                    'Bảo trì': 'Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)'
                })
                
                # Add additional columns
                df_bbnv["Kết quả hoàn thành đánh giá theo phần trăm (%)"] = '100%'
                df_bbnv["Thời gian hoàn thành"] = self.get_time()
                df_bbnv["Kết quả hoàn thành tương ứng nỗ lực xây mới (số MM)"] = 0
                
                # Reorder columns
                df_bbnv = df_bbnv[[
                    "TT", "Nội dung công việc chi tiết", "Thời gian hoàn thành",
                    "Kết quả hoàn thành tương ứng nỗ lực xây mới (số MM)",
                    "Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)",
                    "Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)",
                    "Kết quả hoàn thành đánh giá theo phần trăm (%)"
                ]]
                
                # Apply styling
                df_bbnv.replace(0, numpy.nan, inplace=True)
                df_bbnv = StyleFrame(df_bbnv)
                
                # Set column widths
                column_widths = {
                    "TT": 6.67,
                    "Nội dung công việc chi tiết": 47.67,
                    "Thời gian hoàn thành": 16.50,
                    "Kết quả hoàn thành tương ứng nỗ lực xây mới (số MM)": 14.60,
                    "Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)": 15.83,
                    "Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)": 22.83,
                    "Kết quả hoàn thành đánh giá theo phần trăm (%)": 13.17
                }
                
                for col, width in column_widths.items():
                    df_bbnv.set_column_width(col, width=width)
                
                # Apply fonts and alignment
                font_style = Styler(font="Times New Roman")
                indexes_to_bold = df_bbnv[df_bbnv['TT'].apply(lambda x: not str(x).isnumeric())]
                indexes_not_to_bold = df_bbnv[df_bbnv['TT'].apply(lambda x: str(x).isnumeric())]
                
                df_bbnv.apply_style_by_indexes(
                    cols_to_style='Nội dung công việc chi tiết',
                    indexes_to_style=indexes_to_bold,
                    styler_obj=Styler(horizontal_alignment='left', bold=True, font="Times New Roman"),
                    complement_style=Styler(horizontal_alignment='left', font="Times New Roman")
                )
                
                # Apply right alignment to numeric columns
                col_right = [
                    "Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)",
                    "Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)"
                ]
                
                for col in col_right:
                    df_bbnv.apply_style_by_indexes(
                        cols_to_style=col,
                        indexes_to_style=indexes_not_to_bold,
                        styler_obj=Styler(horizontal_alignment='right', bold=False, font="Times New Roman")
                    )
                    df_bbnv.apply_style_by_indexes(
                        cols_to_style=col,
                        indexes_to_style=indexes_to_bold,
                        styler_obj=Styler(horizontal_alignment='right', bold=True, font="Times New Roman")
                    )
                
                col_num = ['TT', 'Thời gian hoàn thành', 'Kết quả hoàn thành đánh giá theo phần trăm (%)']
                for col in col_num:
                    df_bbnv.apply_style_by_indexes(
                        cols_to_style=col,
                        indexes_to_style=indexes_to_bold,
                        styler_obj=Styler(bold=True, font="Times New Roman"),
                        complement_style=Styler(font="Times New Roman")
                    )
                
                df_bbnv.apply_headers_style(styler_obj=Styler(bold=True, font="Times New Roman"))
                
                # Save to file
                output_path = os.path.join(output_dir, f'{str(hd).replace("/","_")}.xlsx')
                ew = StyleFrame.ExcelWriter(output_path)
                df_bbnv.to_excel(ew)
                ew.close()

            self.status_label.setText("Xử lý hoàn tất!")
            QMessageBox.information(self, "Thành công", f"Các file đã được tạo trong thư mục:\n{output_dir}")

        except Exception as e:
            self.status_label.setText("Đã xảy ra lỗi!")
            QMessageBox.critical(self, "Lỗi", str(e))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_()) 