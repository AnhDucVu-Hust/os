import calendar

import numpy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from styleframe import StyleFrame, Styler, utils
import pandas as pd
import roman
import os
from datetime import datetime,timedelta
import warnings
warnings.filterwarnings("ignore")

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,startcol=0,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:


    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = StyleFrame.ExcelWriter(filename, mode='a',if_sheet_exists='overlay')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol,**to_excel_kwargs)

    # save the workbook
    writer.save()

def find_first_monday(year, month):
    d = datetime(year, int(month), 7)
    offset = -d.weekday() #weekday = 0 means monday
    time = str(d + timedelta(offset))
    y,m,d = str(time.split()[0]).split('-')
    return d+"/"+m+"/"+y
def last_business_day_in_month(year: int, month: int) -> int:
    return str(max(calendar.monthcalendar(year, month)[-1][:5]))+"/"+str(month)+"/"+str(year)

def get_time(year,month):
    return "03/06/2024 " + " đến " +"28/06/2024"
#print(get_time(2023,10))
#def adjust_width()
df_bbnv={}
df=StyleFrame.read_excel("./Task OS tháng 6.xlsx",sheet_name="Sheet0")
hop_dong=list(set(df["Hợp đồng"]))
df_hd={}
df_chuan={}
for hd in hop_dong:
    df_chuan[hd]=pd.DataFrame(columns=['Id','Row label','Bảo trì','Nâng cấp','Total'])
    print(hd)
    df_hd[hd]=df.loc[df["Hợp đồng"]==hd]
    he_thongs=list(set(df_hd[hd]["Hệ thống&CTKT"]))
    id_he_thong=1
    tong_bao_tri=0
    tong_nang_cap=0
    for he_thong in he_thongs:
        df_he_thong= df_hd[hd].loc[df_hd[hd]["Hệ thống&CTKT"]==he_thong]
        stories = list(set(df_he_thong["Tên story"]))
        id_story=1
        bao_tri = df_he_thong.loc[df_he_thong["Phân loại"]=='Bảo trì']["ULNL task"].sum()
        nang_cap = df_he_thong.loc[df_he_thong["Phân loại"]=='Nâng cấp']["ULNL task"].sum()
        total= bao_tri+nang_cap
        tong_nang_cap += nang_cap
        tong_bao_tri += bao_tri
        df_chuan[hd]=df_chuan[hd].append({'Id': roman.toRoman(id_he_thong),'Row label':he_thong,'Bảo trì':bao_tri,"Nâng cấp":nang_cap,'Total':total},ignore_index=True)
        for story in stories:
            df_xet = df_he_thong.loc[df_he_thong["Tên story"]==story]
            bao_tri = df_xet.loc[df_xet["Phân loại"] == 'Bảo trì']["ULNL task"].sum()
            nang_cap = df_xet.loc[df_xet["Phân loại"] == 'Nâng cấp']["ULNL task"].sum()
            df_xet = df_xet[["Hệ thống&CTKT","Tên story","Summary","Phân loại","ULNL task"]]
            total=bao_tri+nang_cap
            df_chuan[hd]=df_chuan[hd].append({'Id':str(roman.toRoman(id_he_thong))+"."+str(id_story),"Row label":story,'Bảo trì':bao_tri,"Nâng cấp":nang_cap,'Total':total},ignore_index=True)
            tasks = list(set(df_xet["Summary"]))
            id_task=1
            for task in tasks:
                df_xet2= df_xet.loc[df_xet["Summary"]==task]
                bao_tri = df_xet2.loc[df_xet2["Phân loại"] == 'Bảo trì']["ULNL task"].sum()
                nang_cap = df_xet2.loc[df_xet2["Phân loại"] == 'Nâng cấp']["ULNL task"].sum()
                #df_xet2 = df_xet2[["Hệ thống & CTKT", "Tên story & [CV]", "Summary task", "Phân loại", "Task"]]
                total = bao_tri + nang_cap
                df_chuan[hd] = df_chuan[hd].append(
                    {'Id': id_task, "Row label": task,
                     'Bảo trì': bao_tri, "Nâng cấp": nang_cap, 'Total': total}, ignore_index=True)
                id_task +=1
            id_story+=1
        id_he_thong +=1
    bao_tri=df_chuan[hd]["Bảo trì"].sum()
    nang_cap = df_chuan[hd]["Nâng cấp"].sum()
    total = df_chuan[hd]["Total"].sum()
    df_chuan[hd] = df_chuan[hd].append({'Id':'',"Row label":'Tổng','Bảo trì':tong_bao_tri,"Nâng cấp":tong_nang_cap,'Total':total},ignore_index=True)
    df_bbnv=df_chuan[hd].rename(columns={'Id':'TT','Row label':'Nội dung công việc chi tiết','Nâng cấp':'Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)','Bảo trì':'Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)'})
    df_bbnv["Kết quả hoàn thành đánh giá theo phần trăm (%)"]='100%'
    df_bbnv["Thời gian hoàn thành"] = get_time(2024, 1)
    df_bbnv["Kết quả hoàn thành tương ứng nỗ lực xây mới (số MM)"] = 0
    df_bbnv = df_bbnv[["TT", "Nội dung công việc chi tiết", "Thời gian hoàn thành",
                       "Kết quả hoàn thành tương ứng nỗ lực xây mới (số MM)",
                       "Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)",
                       "Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)",
                       "Kết quả hoàn thành đánh giá theo phần trăm (%)"]]
    df_bbnv.replace(0,numpy.NAN,inplace=True)
    df_bbnv= StyleFrame(df_bbnv)
    df_bbnv.set_column_width("TT",width=6.67)
    df_bbnv.set_column_width("Nội dung công việc chi tiết", width=47.67)
    df_bbnv.set_column_width("Thời gian hoàn thành", width=16.50)
    df_bbnv.set_column_width("Kết quả hoàn thành tương ứng nỗ lực xây mới (số MM)", width=14.60)
    df_bbnv.set_column_width("Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)", width=15.83)
    df_bbnv.set_column_width("Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)", width=22.83)
    df_bbnv.set_column_width("Kết quả hoàn thành đánh giá theo phần trăm (%)", width=13.17)
    #df_bbnv.style.apply(bold_style, axis=None)
    font_style = Styler(font="Times New Roman")
    indexes_to_bold=df_bbnv[df_bbnv['TT'].apply(lambda x: not str(x).isnumeric())]
    indexes_not_to_bold = df_bbnv[df_bbnv['TT'].apply(lambda x: str(x).isnumeric())]
    print(indexes_to_bold)
    df_bbnv.apply_style_by_indexes(
        cols_to_style='Nội dung công việc chi tiết',indexes_to_style=indexes_to_bold,
        styler_obj=Styler(horizontal_alignment='left',bold=True,font="Times New Roman"),complement_style=Styler(horizontal_alignment='left',font="Times New Roman"))
    col_right=["Kết quả hoàn thành tương ứng nỗ lực nâng cấp (số MM)","Kết quả hoàn thành tương ứng nỗ lực bảo trì (số MM)"]
    #df_bbnv.iloc[list(indexes_to_bold),col_right].style=Styler(horizontal_alignment='right',bold=True, font=utils.fonts.times)
    for col in col_right:
        df_bbnv.apply_style_by_indexes(
                cols_to_style=col,
                indexes_to_style=indexes_not_to_bold,
                styler_obj=Styler(horizontal_alignment='right',bold=False,font="Times New Roman"),
                #complement_style=Styler(horizontal_alignment='right',bold=True,font="Times New Roman"),
            )
        df_bbnv.apply_style_by_indexes(
            cols_to_style=col,
            indexes_to_style=indexes_to_bold,
            styler_obj=Styler(horizontal_alignment='right', bold=True,font="Times New Roman" ),
            #omplement_style=Styler(horizontal_alignment='right', bold=True, font="Times New Roman"),
        )
    col_num=['TT','Thời gian hoàn thành','Kết quả hoàn thành đánh giá theo phần trăm (%)']
    for col in col_num:
        df_bbnv.apply_style_by_indexes(
            cols_to_style=col,
            indexes_to_style=indexes_to_bold,
            styler_obj=Styler(bold=True, font="Times New Roman"),
            complement_style=Styler(font="Times New Roman")
        )
    df_bbnv.apply_headers_style(styler_obj=Styler(bold=True,font="Times New Roman"))
    #df_bbnv = df_bbnv.apply_style_by_indexes(indexes_to_style=df_bbnv.index,
    #                                         styler_obj=font_style)
    ew = StyleFrame.ExcelWriter(f'./my_excel_{str(hd).replace("/","_")}.xlsx')
    '''
    if 'LIFESUP' in str(hd):
        #book = load_workbook("./form LIFESUP.xlsx")
        append_df_to_excel("./form LIFESUP.xlsx",df_bbnv,sheet_name="PGV ",startrow=16)
        #sf_writer = StyleFrame.ExcelWriter("./form LIFESUP.xlsx")
        #sheet_name = 'PGV'
        #df_bbnv.to_excel(sf_writer, sheet_name=sheet_name, startrow=20,
        #            startcol=0)
        #sf_writer.save()
        print("STOP")
        break
    '''
    df_bbnv.to_excel(ew)
    ew.save()


